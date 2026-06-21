"""
Generate competition-format algorithm validation report in Excel.

Usage:
    python utils/report_excel.py --input results.json --output report.xlsx
    python utils/report_excel.py --demo  # generate demo report
"""
import argparse
import json
from pathlib import Path
from typing import Optional

import numpy as np


def generate_validation_report(
    results: dict,
    output_path: str,
    template: bool = False,
):
    """
    Generate an Excel validation report matching competition requirements.

    Parameters
    ----------
    results : dict
        Expected keys (all optional — missing keys become "N/A"):
        - summary: {accuracy, kappa, f1_macro, precision, recall, latency_ms}
        - per_class: [{class_name, accuracy, precision, recall, f1, n_samples}]
        - confusion_matrix: list[list[int]]
        - ablation: [{config, accuracy, kappa, f1_macro}]
        - per_subject: [{subject_id, accuracy, n_trials}]
    output_path : str
    template : bool
        If True, generate a blank template instead.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed. Run: pip install openpyxl")
        return

    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    def style_header(ws, row, n_cols):
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

    def style_data(ws, start_row, end_row, n_cols):
        for row in range(start_row, end_row + 1):
            for col in range(1, n_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = center_align

    # ==================================================================
    # Sheet 1: Summary
    # ==================================================================
    ws1 = wb.active
    ws1.title = "Summary"

    summary = results.get("summary", {})
    headers = ["Metric", "Value"]
    for c, h in enumerate(headers, 1):
        ws1.cell(row=1, column=c, value=h)
    style_header(ws1, 1, len(headers))

    metrics = [
        ("Accuracy", summary.get("accuracy", "N/A")),
        ("Cohen's Kappa", summary.get("kappa", "N/A")),
        ("F1 (macro)", summary.get("f1_macro", "N/A")),
        ("Precision (macro)", summary.get("precision", "N/A")),
        ("Recall (macro)", summary.get("recall", "N/A")),
        ("Decision Latency (ms)", summary.get("latency_ms", "N/A")),
        ("Dataset", summary.get("dataset", "N/A")),
        ("Model", summary.get("model", "N/A")),
        ("N Subjects", summary.get("n_subjects", "N/A")),
        ("N Trials", summary.get("n_trials", "N/A")),
        ("N Channels", summary.get("n_channels", "N/A")),
    ]
    for i, (metric, value) in enumerate(metrics, 2):
        ws1.cell(row=i, column=1, value=metric)
        ws1.cell(row=i, column=2, value=value)
    style_data(ws1, 2, len(metrics) + 1, len(headers))

    # Adjust column widths
    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 20

    # ==================================================================
    # Sheet 2: Per-Class Accuracy
    # ==================================================================
    ws2 = wb.create_sheet("Per-Class")
    headers2 = ["Class", "Accuracy", "Precision", "Recall", "F1", "N Samples"]
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=c, value=h)
    style_header(ws2, 1, len(headers2))

    per_class = results.get("per_class", [])
    if per_class:
        for i, entry in enumerate(per_class, 2):
            ws2.cell(row=i, column=1, value=entry.get("class_name", "N/A"))
            ws2.cell(row=i, column=2, value=entry.get("accuracy", "N/A"))
            ws2.cell(row=i, column=3, value=entry.get("precision", "N/A"))
            ws2.cell(row=i, column=4, value=entry.get("recall", "N/A"))
            ws2.cell(row=i, column=5, value=entry.get("f1", "N/A"))
            ws2.cell(row=i, column=6, value=entry.get("n_samples", "N/A"))
        style_data(ws2, 2, len(per_class) + 1, len(headers2))
    else:
        ws2.cell(row=2, column=1, value="No per-class data available")

    for col_letter in ["A", "B", "C", "D", "E", "F"]:
        ws2.column_dimensions[col_letter].width = 14

    # ==================================================================
    # Sheet 3: Confusion Matrix
    # ==================================================================
    ws3 = wb.create_sheet("Confusion Matrix")
    cm = results.get("confusion_matrix", [])
    if cm:
        n = len(cm)
        # Header row
        ws3.cell(row=1, column=1, value="True \\ Pred")
        for j in range(n):
            ws3.cell(row=1, column=j + 2, value=f"Class {j}")
        style_header(ws3, 1, n + 1)

        for i, row in enumerate(cm):
            ws3.cell(row=i + 2, column=1, value=f"Class {i}")
            ws3.cell(row=i + 2, column=1).font = Font(bold=True)
            for j, val in enumerate(row):
                ws3.cell(row=i + 2, column=j + 2, value=val)
        style_data(ws3, 2, n + 1, n + 1)
    else:
        ws3.cell(row=1, column=1, value="No confusion matrix data available")

    # ==================================================================
    # Sheet 4: Ablation Comparison
    # ==================================================================
    ws4 = wb.create_sheet("Ablation")
    headers4 = ["Configuration", "Accuracy", "Kappa", "F1 (macro)"]
    for c, h in enumerate(headers4, 1):
        ws4.cell(row=1, column=c, value=h)
    style_header(ws4, 1, len(headers4))

    ablation = results.get("ablation", [])
    if ablation:
        for i, entry in enumerate(ablation, 2):
            ws4.cell(row=i, column=1, value=entry.get("config", "N/A"))
            ws4.cell(row=i, column=2, value=entry.get("accuracy", "N/A"))
            ws4.cell(row=i, column=3, value=entry.get("kappa", "N/A"))
            ws4.cell(row=i, column=4, value=entry.get("f1_macro", "N/A"))
        style_data(ws4, 2, len(ablation) + 1, len(headers4))
    else:
        ws4.cell(row=2, column=1, value="No ablation data available")

    ws4.column_dimensions["A"].width = 40
    for col_letter in ["B", "C", "D"]:
        ws4.column_dimensions[col_letter].width = 16

    # ==================================================================
    # Sheet 5: Per-Subject
    # ==================================================================
    ws5 = wb.create_sheet("Per-Subject")
    headers5 = ["Subject ID", "Accuracy", "N Trials"]
    for c, h in enumerate(headers5, 1):
        ws5.cell(row=1, column=c, value=h)
    style_header(ws5, 1, len(headers5))

    per_subject = results.get("per_subject", [])
    if per_subject:
        for i, entry in enumerate(per_subject, 2):
            ws5.cell(row=i, column=1, value=entry.get("subject_id", "N/A"))
            ws5.cell(row=i, column=2, value=entry.get("accuracy", "N/A"))
            ws5.cell(row=i, column=3, value=entry.get("n_trials", "N/A"))
        style_data(ws5, 2, len(per_subject) + 1, len(headers5))
    else:
        ws5.cell(row=2, column=1, value="No per-subject data available")

    for col_letter in ["A", "B", "C"]:
        ws5.column_dimensions[col_letter].width = 18

    # Save
    wb.save(output_path)
    print(f"Report saved to {output_path}")


def generate_demo_report(output_path: str = "validation_report_demo.xlsx"):
    """Generate a demo report with placeholder values to show the format."""
    results = {
        "summary": {
            "accuracy": 0.753,
            "kappa": 0.682,
            "f1_macro": 0.741,
            "precision": 0.748,
            "recall": 0.739,
            "latency_ms": 45,
            "dataset": "PhysioNet MI + Self-collected (n=20)",
            "model": "EEGNet + SpatiotemporalAttention",
            "n_subjects": 30,
            "n_trials": 5400,
            "n_channels": 16,
        },
        "per_class": [
            {"class_name": "Idle (0)", "accuracy": 0.82, "precision": 0.79, "recall": 0.85, "f1": 0.82, "n_samples": 1800},
            {"class_name": "Left MI (1)", "accuracy": 0.71, "precision": 0.73, "recall": 0.68, "f1": 0.70, "n_samples": 1800},
            {"class_name": "Right MI (2)", "accuracy": 0.72, "precision": 0.70, "recall": 0.73, "f1": 0.71, "n_samples": 1800},
        ],
        "confusion_matrix": [
            [1476, 162, 162],
            [288, 1278, 234],
            [252, 234, 1314],
        ],
        "ablation": [
            {"config": "EEGNet (base)", "accuracy": 0.527, "kappa": 0.291, "f1_macro": 0.518},
            {"config": "EEGNet + SE-Attn", "accuracy": 0.568, "kappa": 0.352, "f1_macro": 0.559},
            {"config": "EEGNet + MHSA-Attn", "accuracy": 0.614, "kappa": 0.421, "f1_macro": 0.607},
            {"config": "EEGNet + Temporal-Attn", "accuracy": 0.582, "kappa": 0.373, "f1_macro": 0.574},
            {"config": "EEGNet + Spatiotemporal-Attn", "accuracy": 0.637, "kappa": 0.456, "f1_macro": 0.631},
        ],
        "per_subject": [
            {"subject_id": f"S{i:02d}", "accuracy": round(0.65 + np.random.RandomState(i).rand() * 0.25, 3), "n_trials": 180}
            for i in range(1, 31)
        ],
    }
    generate_validation_report(results, output_path)


def main():
    parser = argparse.ArgumentParser(description="Generate validation report Excel")
    parser.add_argument("--input", help="JSON file with results dict")
    parser.add_argument("--output", default="validation_report.xlsx")
    parser.add_argument("--demo", action="store_true", help="Generate demo report")
    args = parser.parse_args()

    if args.demo:
        generate_demo_report(args.output)
    elif args.input:
        with open(args.input) as f:
            results = json.load(f)
        generate_validation_report(results, args.output)
    else:
        print("Usage: python utils/report_excel.py --input results.json")
        print("       python utils/report_excel.py --demo")


if __name__ == "__main__":
    main()
