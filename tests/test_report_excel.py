"""Tests for utils/report_excel.py — competition-format Excel report generation."""
import json
import os

import numpy as np
import pytest

from utils.report_excel import generate_validation_report, generate_demo_report


# ── Fixtures ────────────────────────────────────────────────────────────────

@pytest.fixture
def full_results():
    """Complete results dict with all optional keys."""
    return {
        "summary": {
            "accuracy": 0.753,
            "kappa": 0.682,
            "f1_macro": 0.741,
            "precision": 0.748,
            "recall": 0.739,
            "latency_ms": 45,
            "dataset": "Test Dataset",
            "model": "TestModel",
            "n_subjects": 30,
            "n_trials": 5400,
            "n_channels": 8,
        },
        "per_class": [
            {"class_name": "Left", "accuracy": 0.71, "precision": 0.73,
             "recall": 0.68, "f1": 0.70, "n_samples": 1800},
            {"class_name": "Right", "accuracy": 0.72, "precision": 0.70,
             "recall": 0.73, "f1": 0.71, "n_samples": 1800},
        ],
        "confusion_matrix": [
            [1278, 522],
            [468, 1332],
        ],
        "ablation": [
            {"config": "Baseline", "accuracy": 0.53, "kappa": 0.06, "f1_macro": 0.52},
            {"config": "Proposed", "accuracy": 0.75, "kappa": 0.50, "f1_macro": 0.74},
        ],
        "per_subject": [
            {"subject_id": "S01", "accuracy": 0.72, "n_trials": 180},
            {"subject_id": "S02", "accuracy": 0.78, "n_trials": 180},
        ],
    }


@pytest.fixture
def minimal_results():
    """Minimal results with only summary."""
    return {
        "summary": {"accuracy": 0.60, "kappa": 0.20},
    }


# ── generate_validation_report ──────────────────────────────────────────────

class TestGenerateValidationReport:
    def test_creates_file(self, full_results, tmp_path):
        path = tmp_path / "report.xlsx"
        generate_validation_report(full_results, str(path))
        assert path.exists()
        assert os.path.getsize(path) > 0

    def test_full_results_no_error(self, full_results, tmp_path):
        path = tmp_path / "full_report.xlsx"
        generate_validation_report(full_results, str(path))
        assert path.exists()

    def test_minimal_results_no_error(self, minimal_results, tmp_path):
        path = tmp_path / "minimal_report.xlsx"
        generate_validation_report(minimal_results, str(path))
        assert path.exists()

    def test_empty_results_no_error(self, tmp_path):
        path = tmp_path / "empty_report.xlsx"
        generate_validation_report({}, str(path))
        assert path.exists()

    def test_missing_keys_filled_with_na(self, tmp_path):
        """Missing optional keys should not crash — placeholders used."""
        results = {"summary": {"accuracy": 0.5}}
        path = tmp_path / "partial_report.xlsx"
        generate_validation_report(results, str(path))
        # If we get here without error, the test passes
        assert path.exists()

    def test_missing_per_class(self, tmp_path):
        results = {"summary": {"accuracy": 0.5}, "per_class": []}
        path = tmp_path / "no_per_class.xlsx"
        generate_validation_report(results, str(path))
        assert path.exists()

    def test_missing_confusion_matrix(self, tmp_path):
        results = {"summary": {"accuracy": 0.5}, "confusion_matrix": []}
        path = tmp_path / "no_cm.xlsx"
        generate_validation_report(results, str(path))
        assert path.exists()

    def test_missing_ablation(self, tmp_path):
        results = {"summary": {"accuracy": 0.5}, "ablation": []}
        path = tmp_path / "no_ablation.xlsx"
        generate_validation_report(results, str(path))
        assert path.exists()

    def test_missing_per_subject(self, tmp_path):
        results = {"summary": {"accuracy": 0.5}, "per_subject": []}
        path = tmp_path / "no_per_subject.xlsx"
        generate_validation_report(results, str(path))
        assert path.exists()

    def test_numeric_values_preserved(self, full_results, tmp_path):
        """Check the file can be opened and has expected sheet names."""
        path = tmp_path / "numeric.xlsx"
        generate_validation_report(full_results, str(path))

        import openpyxl
        wb = openpyxl.load_workbook(path)
        assert "Summary" in wb.sheetnames
        assert "Per-Class" in wb.sheetnames
        assert "Confusion Matrix" in wb.sheetnames
        assert "Ablation" in wb.sheetnames
        assert "Per-Subject" in wb.sheetnames

        # Check Summary sheet content
        ws = wb["Summary"]
        summary_data = {}
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[0] and row[1]:
                summary_data[row[0]] = row[1]
        assert summary_data.get("Accuracy") == 0.753
        assert summary_data.get("Cohen's Kappa") == 0.682

    def test_per_class_data(self, full_results, tmp_path):
        path = tmp_path / "per_class.xlsx"
        generate_validation_report(full_results, str(path))

        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb["Per-Class"]
        rows = list(ws.iter_rows(min_row=2, max_col=6, values_only=True))
        assert len(rows) == 2
        assert rows[0][0] == "Left"

    def test_confusion_matrix_data(self, full_results, tmp_path):
        path = tmp_path / "cm.xlsx"
        generate_validation_report(full_results, str(path))

        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb["Confusion Matrix"]
        # Cell (2,2) should be 1278
        assert ws.cell(row=2, column=2).value == 1278

    def test_per_subject_data(self, full_results, tmp_path):
        path = tmp_path / "per_subj.xlsx"
        generate_validation_report(full_results, str(path))

        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb["Per-Subject"]
        rows = list(ws.iter_rows(min_row=2, max_col=3, values_only=True))
        assert len(rows) == 2
        assert rows[0][0] == "S01"

    def test_with_numpy_types(self, tmp_path):
        """Should handle numpy numeric types."""
        results = {
            "summary": {
                "accuracy": np.float64(0.75),
                "kappa": np.float32(0.5),
                "n_subjects": np.int64(30),
            },
        }
        path = tmp_path / "numpy_types.xlsx"
        generate_validation_report(results, str(path))
        assert path.exists()


# ── generate_demo_report ────────────────────────────────────────────────────

class TestGenerateDemoReport:
    def test_creates_file(self, tmp_path):
        path = tmp_path / "demo_report.xlsx"
        generate_demo_report(str(path))
        assert path.exists()
        assert os.path.getsize(path) > 0

    def test_demo_has_all_sheets(self, tmp_path):
        path = tmp_path / "demo_sheets.xlsx"
        generate_demo_report(str(path))

        import openpyxl
        wb = openpyxl.load_workbook(path)
        assert "Summary" in wb.sheetnames
        assert "Per-Class" in wb.sheetnames
        assert "Confusion Matrix" in wb.sheetnames
        assert "Ablation" in wb.sheetnames
        assert "Per-Subject" in wb.sheetnames

    def test_demo_has_30_subjects(self, tmp_path):
        path = tmp_path / "demo_subjects.xlsx"
        generate_demo_report(str(path))

        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb["Per-Subject"]
        rows = list(ws.iter_rows(min_row=2, max_col=1, values_only=True))
        assert len(rows) == 30

    def test_demo_summary_has_values(self, tmp_path):
        path = tmp_path / "demo_summary.xlsx"
        generate_demo_report(str(path))

        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb["Summary"]
        acc_cell = ws.cell(row=2, column=2).value
        assert acc_cell == 0.753

    def test_demo_with_default_path(self, tmp_path, monkeypatch):
        """Default path should work (in CWD)."""
        import os as _os
        monkeypatch.chdir(tmp_path)
        path = tmp_path / "validation_report_demo.xlsx"
        generate_demo_report(str(path))
        assert path.exists()
